import pandas as pd
import math
from textblob import TextBlob
from gensim import corpora, models
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.decomposition import LatentDirichletAllocation
import openpyxl
from openpyxl.styles import PatternFill, Font, Color
import os
# Load original CSV file
df = pd.read_excel('./medallia.csv')
# Print column names
print(df.columns)
# Clean up DataFrame
df = df[['Response Id', 'DIS Verbatim', 'Combined DIS', 'Primary Task']]
df = df.dropna(subset=['DIS Verbatim'])
df = df.reset_index(drop=True)
# Rename 'Combined DIS' column to 'DIS Rating'
df.rename(columns={'Combined DIS': 'DIS Rating'}, inplace=True)
# Add sentiment column
def get_sentiment(rating):
    if rating == 5:
        return 'positive'
    elif rating == 4:
        return 'neutral'
    else:
        return 'negative'
df['sentiment'] = df['DIS Rating'].apply(get_sentiment)
# Add category column using Gensim
texts = [text.split() if isinstance(text, str) else [] for text in df['DIS Verbatim']]
dictionary = corpora.Dictionary(texts)
corpus = [dictionary.doc2bow(text) for text in texts]
lda_model = models.LdaModel(corpus, num_topics=5, id2word=dictionary)
df['category'] = [max(lda_model[text], key=lambda x: x[1])[0] for text in corpus]
# Add actionable issues column using scikit-learn
vectorizer = CountVectorizer(stop_words='english')
doc_term_matrix = vectorizer.fit_transform(df['DIS Verbatim'].values.astype('U'))
lda = LatentDirichletAllocation(n_components=5, random_state=0)
lda.fit(doc_term_matrix)
feature_names = vectorizer.get_feature_names_out()
# Get topic distribution for each document
doc_topic_distributions = lda.transform(doc_term_matrix)
# Get the topic number that has the highest percentage contribution in that document
dominant_topic_per_doc = doc_topic_distributions.argmax(axis=1)
# Get top 10 words for each topic
top_words_per_topic = [' '.join([feature_names[i] for i in topic.argsort()[:-11:-1] if feature_names[i].isalpha()]) for topic in lda.components_]
# Map dominant topic to top 10 words of that topic
df['actionable_issues'] = [top_words_per_topic[i] for i in dominant_topic_per_doc]
# Save modified DataFrame to new CSV file
df.to_excel('modified_data.xlsx', index=False)
# Apply styles to the Excel file
workbook = openpyxl.load_workbook('modified_data.xlsx')
worksheet = workbook.active
# Apply fill
greyFill = PatternFill(start_color='00C0C0C0',
                   end_color='00C0C0C0',
                   fill_type='solid')
for cell in worksheet[1]:
    cell.fill = greyFill
    cell.font = Font(bold=True)
workbook.save('modified_data.xlsx')