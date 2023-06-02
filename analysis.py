import os
import pandas as pd
from sklearn.decomposition import LatentDirichletAllocation
from sklearn.feature_extraction.text import CountVectorizer
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import xmltodict

def process_uploaded_file(file_path):
    file_extension = os.path.splitext(file_path)[1]
    if file_extension == '.xlsx':
        sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
        df = pd.concat(sheets.values())
    elif file_extension == '.xls':
        sheets = pd.read_excel(file_path, sheet_name=None, engine='xlrd')
        df = pd.concat(sheets.values())
    elif file_extension == '.csv':
        df = pd.read_csv(file_path)
    elif file_extension == '.json':
        df = pd.read_json(file_path)
    elif file_extension == '.xml':
        with open(file_path, 'r') as f:
            xml_string = f.read()
        xml_data = xmltodict.parse(xml_string)
        df = pd.json_normalize(xml_data['root']['data'])
    else:
        raise ValueError("The file must be an .xls, .xlsx, .csv, .json, or .xml file.")

    # Data preprocessing and analysis code goes here

    workbook = Workbook()
    for df, sheetname in zip([df, df[['Response ID', 'DIS Verbatim', 'DIS Rating', 'Primary Task', 'category', 'actionable_issues']]], ['Original', 'Processed']):
        worksheet = workbook.create_sheet(sheetname)
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)

    # Apply formatting to the workbook

    output_file_path = os.path.join('uploads', 'processed_' + os.path.basename(file_path))
    workbook.save(output_file_path)

    return output_file_path
